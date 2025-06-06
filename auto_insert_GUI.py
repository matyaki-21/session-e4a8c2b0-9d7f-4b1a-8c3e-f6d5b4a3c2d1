from typing import Any, List
import openpyxl
import glob
import os
import xlrd
import fnmatch
import logging
import re
import customtkinter as ctk
from tkinter import messagebox
from sqlalchemy import insert, delete, select
from database import SessionLocal, engine
from models import PurchaseCostVariance, CostCenterReports, ActualExpenses, SalesReceivables, Discounts, StaffingSummary, SalesBacklog, ConstructionSuspenseAccounts, ItemizedInventoryDetails, DirectExpenses, PurchasePriceGap, ItemList, FinancialCostOfSales, FinancialOperatingIncome
from column_insert import insert_year_month_column
from bulk_insert_utils import bulk_insert_with_chunk

# CustomTkinterのテーマを設定
ctk.set_appearance_mode("System")  # "Light" or "Dark" も指定可能
ctk.set_default_color_theme("blue")  # テーマカラーを設定

#################初期値指定エリア#####################
#グローバル変数の定義
SalesBacklog_filename = None
CostCenterReports_filename = None
ActualExpenses_filename = None
SalesReceivables_filename = None
Discounts_filename = None
Staffingtable_filename = None
ConstructionSuspenseAccounts_filename = None
CostVarianceAdjustment_filename = None
ItemizedInventoryDetails_filename = None
DirectExpenses_filename = None
PurchasePriceGap_filename = None
ItemList_filename = None
FinancialCostOfSales_filename = None

def validate_year_month_format(year_month):
    """
    YYYYMM形式の文字列かどうかを検証する
    """
    pattern = r"^\d{4}(0[1-9]|1[0-2])$"  # YYYYは4桁の数字、MMは01から12の間
    return re.match(pattern, year_month) is not None

def check_year_month_format():
    """
    YYYYMM形式でない場合、エラーポップアップを表示する。
    正しい場合はTrue、エラーの場合はFalseを返す。
    """
    year_month = entry_year_month.get()

    if not validate_year_month_format(year_month):
        messagebox.showerror("エラー", "入力はYYYYMM形式である必要があります。")
        return False

    return True

def process_year_month():
    global progress_bar
    global SalesBacklog_filename, CostCenterReports_filename, ActualExpenses_filename
    global SalesReceivables_filename, Discounts_filename, Staffingtable_filename
    global ConstructionSuspenseAccounts_filename, CostVarianceAdjustment_filename
    global ItemizedInventoryDetails_filename, DirectExpenses_filename
    global PurchasePriceGap_filename,ItemList_filename,FinancialCostOfSales_filename

    def process_itemized_inventory_details(input_excel_path, processing_year_month):
        """
        必要なデータを抽出して「品目別在庫明細」テーブルに挿入する関数
        """
        logging.info(f"ファイル {input_excel_path} の読み込みを開始します。")

        try:
            input_wb = openpyxl.load_workbook(input_excel_path, data_only=True)
            input_sheet = input_wb.active
            logging.info(f"ファイル {input_excel_path} の読み込みが完了しました。")
        except Exception as e:
            logging.error(f"ファイル {input_excel_path} の読み込み中にエラーが発生しました: {e}")
            return

        data_list = []

        segments_and_rows = [
            ('D(N0+N1)', 8), ('D(N2)', 9), ('K', 10), ('S', 11), 
            ('BD', 12), ('PSD', 13), ('F', 14)
        ]

        logging.info(f"データ抽出を開始します。")
        for segment, row_index in segments_and_rows:
            for 科目, 科目コード, 金額区分, col in [
                ('購入品', '3040', '在庫額', 'B'), 
                ('購入品', '-', '評価減', 'C'), 
                ('購入品', '-', '差額振分け', 'D'),
                ('購入品', '-', '未着品', 'F'),
                ('製品', '7920', '在庫額', 'H'),
                ('製品', '-', '評価減', 'I'),
                ('製品', '-', '差額振分け', 'J'),
                ('仕掛', '7900', '在庫額', 'L'),
                ('仕掛', '4000', '在庫額', 'M'),
                ('仕掛', '16310', '在庫額', 'N'),
                ('仕掛', '-', '評価減', 'O'),
                ('仕掛', '-', '差額振分け', 'Q'),
            ]:
                amount = input_sheet[f"{col}{row_index}"].value
                data_list.append({
                    '年月': processing_year_month,
                    'セグメント': segment,
                    '科目': 科目,
                    '科目コード': 科目コード,
                    '金額区分': 金額区分,
                    '金額': amount * 1000 if amount else 0
                })
        logging.info(f"データ抽出が完了しました。")

        logging.info(f"データベースへのデータ挿入を開始します。")
        db = SessionLocal()
        try:
            engine.echo = False
            bulk_insert_with_chunk(db, ItemizedInventoryDetails, data_list, chunk_size=1000)
            logging.info(f"データベースへのデータ挿入が完了しました。")
        except Exception as e:
            logging.error(f"データ挿入中にエラーが発生しました: {e}")
        finally:
            db.close()
            engine.echo = True
            logging.info(f"データベース接続を終了しました。")
    
    def process_cost_variance_adjustment(input_excel_path, processing_year_month):
        """
        「原価差額調整計算表」のExcelファイルを読み込み、SQL Serverにデータを挿入する関数
        """
        logging.info(f"ファイル {input_excel_path} の読み込みを開始します。")
        
        input_wb = openpyxl.load_workbook(input_excel_path, data_only=True)
        input_sheet = input_wb.active

        month_column_map = {
            1: 4,  2: 5,  3: 6,  4: 7,  5: 8,  6: 9,
            7: 11, 8: 12, 9: 13, 10: 14, 11: 15, 12: 16
        }
        processing_month = processing_year_month % 100
        target_column_index = month_column_map.get(processing_month)

        if target_column_index is None:
            logging.error(f"無効な月です: {processing_month}")
            return

        data_list = []
        cost_variance_subjects = [
            "材料・購入品在庫_繰越", "材料・購入品在庫_増加", "材料・購入品在庫_減少", "材料・購入品在庫_残高",
            "購入価格差異_繰越", "購入価格差異_増加", "購入価格差異_仕掛品振替", "購入価格差異_廃却分", "購入価格差異_残高",
            "製造原価差額_間接費", "製造原価差額_加工費",
            "調整原差_繰越", "調整原差_間接費差異", "調整原差_購入価額差異", "調整原差_標準原価評価替差額", "調整原差_その他原価差額",
            "仕掛品在庫_繰越", "仕掛品在庫_増加", "仕掛品在庫_他勘定", "仕掛品在庫_製品振替", "仕掛品在庫_残高",
            "製品在庫_繰越", "製品在庫_増加", "製品在庫_他勘定", "製品在庫_売上原価振替", "製品在庫_残高",
            "原価差額負担額"
        ]
        target_rows = [5, 6, 7, 8, 13, 14, 15, 16, 17, 22, 23,
                    29, 30, 31, 32, 33, 40, 41, 42, 43, 44,
                    49, 50, 51, 52, 53, 65]

        for subject, row_index in zip(cost_variance_subjects, target_rows):
            amount = input_sheet.cell(row=row_index, column=target_column_index).value

            data_list.append({
                '年月': processing_year_month,
                'セグメント': 'ALL',
                '購入原価差額科目': subject,
                '金額': amount if amount is not None else 0
            })

        logging.info(f"データ抽出が完了しました。")
        
        logging.info(f"データベースへのデータ挿入を開始します。")
        db = SessionLocal()
        try:
            engine.echo = False
            bulk_insert_with_chunk(db, PurchaseCostVariance, data_list, chunk_size=1000)
            logging.info(f"データベースへのデータ挿入が完了しました。")
        except Exception as e:
            logging.error(f"データ挿入中にエラーが発生しました：{e}")
        finally:
            db.close()
            engine.echo = True
            logging.info(f"データベース接続を終了しました。")
                
    def process_construction_suspense_accounts(input_excel_path, processing_year_month):
        """
        「建仮計上額」のExcelファイルを読み込み、SQL Serverにデータを挿入する関数
        """
        logging.info(f"ファイル {input_excel_path} の読み込みを開始します。")
        
        input_wb = openpyxl.load_workbook(input_excel_path, data_only=True)
        sheet_name = str(processing_year_month // 100) + "年度実績"
        input_sheet = input_wb[sheet_name]

        target_column_index = processing_year_month % 100 + 1

        logging.info(f"データ抽出を開始します。")
        data_list = []
        for row_index in range(2, 5):  # 2行目から4行目まで
            segment = 'D' if row_index == 2 else ('PD' if row_index == 3 else 'F')
            amount = input_sheet.cell(row=row_index, column=target_column_index).value

            data_list.append({
                '年月': processing_year_month,
                'セグメント': segment,
                '金額': amount if amount is not None else 0
            })
        logging.info(f"データ抽出が完了しました。")

        logging.info(f"データベースへのデータ挿入を開始します。")
        db = SessionLocal()
        try:
            engine.echo = False
            bulk_insert_with_chunk(db, ConstructionSuspenseAccounts, data_list, chunk_size=1000)
            logging.info(f"データベースへのデータ挿入が完了しました。")
        except Exception as e:
            logging.error(f"データ挿入中にエラーが発生しました：{e}")
        finally:
            db.close()
            engine.echo = True
            logging.info(f"データベース接続を終了しました。")

    def process_direct_expenses(input_excel_path, processing_year_month):
        """
        「PJ管理物件勘定内訳表」のExcelファイルを読み込み、指定されたデータをSQL Serverに挿入する関数
        """
        logging.info(f"ファイル {input_excel_path} の読み込みを開始します。")

        try:
            input_wb = openpyxl.load_workbook(input_excel_path, data_only=True)
            input_sheet = input_wb.active
            logging.info(f"ファイル {input_excel_path} の読み込みが完了しました。")
        except Exception as e:
            logging.error(f"ファイル {input_excel_path} の読み込み中にエラーが発生しました: {e}")
            return

        month = processing_year_month % 100
        row_mapping = {
            1: 346, 2: 349, 3: 352, 4: 355, 5: 358, 6: 361,
            7: 346, 8: 349, 9: 352, 10: 355, 11: 358, 12: 361
        }

        row_index = row_mapping.get(month)
        if not row_index:
            logging.error(f"無効な年月: {processing_year_month}")
            return

        amount = input_sheet[f"AA{row_index}"].value
        if amount is None:
            logging.error(f"行 {row_index} でデータが見つかりません")
            return

        data_list = [{
            '年月': processing_year_month,
            'SBU': 'PD',
            '金額': amount
        }]

        logging.info(f"データベースへのデータ挿入を開始します。")
        db = SessionLocal()
        try:
            engine.echo = False
            bulk_insert_with_chunk(db, DirectExpenses, data_list, chunk_size=1000)
            logging.info(f"データベースへのデータ挿入が完了しました。")
        except Exception as e:
            logging.error(f"データ挿入中にエラーが発生しました: {e}")
        finally:
            db.close()
            engine.echo = True
            logging.info(f"データベース接続を終了しました。")

    def process_financialcostOfsales(input_excel_path, processing_year_month):
        """
        「財務諸表」のExcelファイルを読み込み、指定されたデータ（年月, 金額）をSQL Serverに挿入する関数
        """
        logging.info(f"ファイル {input_excel_path} の読み込みを開始します。")
        try:
            input_wb = openpyxl.load_workbook(input_excel_path, data_only=True)
            input_sheet = input_wb.active
            logging.info(f"ファイル {input_excel_path} の読み込みが完了しました。")
        except Exception as e:
            logging.error(f"ファイル {input_excel_path} の読み込み中にエラーが発生しました: {e}")
            return

        target_strings = ["売上原価  合計", "営業利益"]
        records_to_insert = []

        for target_string in target_strings:
            target_row = None
            for row in input_sheet.iter_rows(min_row=1):
                cell_b = row[1]  # B列
                if cell_b.value == target_string:
                    target_row = row
                    break

            if target_row is None:
                logging.error(f"Excelファイル内に B列が '{target_string}' の行が見つかりません。")
                continue

            amount = target_row[7].value  # H列（8列目）
            if amount is None:
                logging.error(f"{target_string} の H列 に金額がありません。")
                continue

            if isinstance(amount, str):
                try:
                    amount = float(amount.replace(",", "").strip())
                except ValueError:
                    logging.error(f"{target_string} の金額変換に失敗しました: {amount}")
                    continue

            elif not isinstance(amount, (int, float)):
                logging.error(f"{target_string} の金額が不正な形式です: {amount}")
                continue

            if target_string == "売上原価  合計":
                records_to_insert.append({
                    '年月': processing_year_month,
                    '金額': amount,
                    '_model': FinancialCostOfSales
                })
            elif target_string == "営業利益":
                records_to_insert.append({
                    '年月': processing_year_month,
                    '金額': amount,
                    '_model': FinancialOperatingIncome
                })

        db = SessionLocal()
        try:
            engine.echo = False
            for entry in records_to_insert:
                model = entry.pop('_model')
                # すでに存在するレコードがある場合は delete してから insert でも可
                db.query(model).filter_by(年月=entry['年月']).delete()
                bulk_insert_with_chunk(db, model, [entry], chunk_size=1000)
            logging.info("財務諸表データのバルク挿入が完了しました。")
        except Exception as e:
            db.rollback()
            logging.error(f"財務諸表のデータ挿入中にエラーが発生しました: {e}")
        finally:
            db.close()
            engine.echo = True
            logging.info("データベース接続を終了しました。")

    if not check_year_month_format():
        return  # YYYYMM形式でない場合は処理を中断
    
    year_month = entry_year_month.get()
    
    progress_bar.set(0)
    progress_bar.grid(row=4, column=0, columnspan=2, padx=20, pady=20)  # プログレスバーを表示

    # YYYYMM形式から年と月を取得
    year = year_month[:4]  # 年の部分を取得
    month = int(year_month[4:])  # 月の部分を整数で取得

    # 年の後ろ2桁を取得
    year_last_two_digits = year[2:]

    # 月が7月未満かどうかでDirectExpenses_sheetnameを決定
    if month < 7:
        DirectExpenses_sheetname = f'{year_last_two_digits}・6'
    else:
        DirectExpenses_sheetname = f'{year_last_two_digits}・12'
    
    sheet_name_mapping = {
        '受注売上受注残*': str(year),
        '*建仮計上額.xls': f'{year}年度実績',
        '人員集計表*': f'{year_last_two_digits}年度実績',
        '*品目別在庫明細表.xlsx': '差額有',
        'PJ管理物件勘定内訳表*': DirectExpenses_sheetname,
        '*原価差額調整計算表 AI.xlsx': '年度調整版'
    }

    def get_sheet_name(filename):
        """ ファイル名に基づいてシート名を返す """
        for pattern, sheet_name in sheet_name_mapping.items():
            if fnmatch.fnmatch(filename, pattern):
                return sheet_name
        return None  # 指定がない場合

    try:
        processing_year_month = int(entry_year_month.get())

        #input_dir_path = 'C:/Users/1221050/Documents/work/01_業務/BI/Board/総務課/データソース/' + str(processing_year_month)
        input_dir_path = 'W:/WF_07_ｼｽﾃﾑ関係/16_BI/総務課/データソース/' + str(processing_year_month)
        #input_dir_path = 'C:/Users/KYOHEI/Documents/work/データソース/' + str(processing_year_month)
        output_dir_path = os.path.join(input_dir_path, 'TMP')
        
        logging.basicConfig(
            filename=input_dir_path + '/app.log',
            level=logging.DEBUG,
            format='%(asctime)s - %(levelname)s - %(message)s',
            filemode='a'
        )
        
        insert_year_month_column(input_dir_path, output_dir_path, processing_year_month, get_sheet_name)
        
        files = os.listdir(output_dir_path)
        total_files = len(files)  # ファイルの総数
        
        if total_files == 0:
            messagebox.showinfo("情報", "処理対象のファイルがありません。")
            progress_bar.grid_remove()
            return

        search_patterns = [
            ('受注売上受注残*', 'SalesBacklog_filename'),
            ('売上実績*', 'CostCenterReports_filename'),
            ('S4データ*', 'ActualExpenses_filename'),
            ('*SBU別売上債権（各種売掛金）*', 'SalesReceivables_filename'),
            ('*値引き.xlsx', 'Discounts_filename'),
            ('人員集計表*', 'Staffingtable_filename'),
            ('*建仮計上額.xlsx', 'ConstructionSuspenseAccounts_filename'),
            ('*原価差額調整計算表 AI.xlsx', 'CostVarianceAdjustment_filename'),
            ('*品目別在庫明細表.xlsx', 'ItemizedInventoryDetails_filename'),
            ('PJ管理物件勘定内訳表*', 'DirectExpenses_filename'),
            ('*AQZZCO_CT.xlsx', 'PurchasePriceGap_filename'),
            ('品目一覧表.xlsx', 'ItemList_filename'),
            ('財務諸表*', 'FinancialCostOfSales_filename')
        ]

        for search_string, variable_name in search_patterns:
            file_list = glob.glob(output_dir_path + '/' + search_string)
            if file_list:
                globals()[variable_name] = os.path.basename(file_list[0])
                logging.info(f'{variable_name} に {os.path.basename(file_list[0])} を格納しました。')
        
        # モデルとExcelファイルのマッピング
        model_mapping = {
            CostCenterReports_filename: CostCenterReports,
            ActualExpenses_filename: ActualExpenses,
            SalesReceivables_filename: SalesReceivables,
            SalesBacklog_filename: SalesBacklog,
            Discounts_filename: Discounts,
            PurchasePriceGap_filename: PurchasePriceGap,
            ItemList_filename: ItemList,
            FinancialCostOfSales_filename: FinancialCostOfSales
        }
        
        # Excelファイルごとの開始行と開始列を指定する辞書
        start_config_mapping = {
            CostCenterReports_filename: {'start_row': 2, 'start_column': 1},
            ActualExpenses_filename: {'start_row': 2, 'start_column': 1},
            SalesReceivables_filename: {'start_row': 2, 'start_column': 1},
            Discounts_filename: {'start_row': 2, 'start_column': 1},
            SalesBacklog_filename: {'start_row': 6, 'start_column': 1},
            Staffingtable_filename: {'start_row': 59, 'start_column': 8},
            PurchasePriceGap_filename: {'start_row': 2, 'start_column': 1},
            ItemList_filename: {'start_row': 2, 'start_column': 1},
            FinancialCostOfSales_filename: {'start_row': 2, 'start_column': 1}
        }

        # ディレクトリ内のExcelファイルを処理
        for index, filename in enumerate(files):
            if filename.endswith('.xlsx') or filename.endswith('.xls'):
                input_excel_path = os.path.join(output_dir_path, filename)

                # ファイルの拡張子に応じて適切なライブラリを使用
                if filename.endswith('.xlsx'):
                    input_wb = openpyxl.load_workbook(input_excel_path, data_only=True)  
                    input_sheet = input_wb.active
                else:
                    input_wb = xlrd.open_workbook(input_excel_path)
                    input_sheet = input_wb.sheet_by_index(0)

                # モデルを取得
                model = model_mapping.get(filename)
                if model is None and filename not in (Staffingtable_filename, ConstructionSuspenseAccounts_filename, CostVarianceAdjustment_filename, ItemizedInventoryDetails_filename, DirectExpenses_filename, FinancialCostOfSales_filename):
                    print(f"対応するモデルが見つかりません：{filename}")
                    messagebox.showinfo("エラー", "モデル取得でエラーが発生しました。ログを確認してください。")
                    continue

                # 開始行、開始列を取得
                start_config = start_config_mapping.get(filename, {'start_row': 2, 'start_column': 1})
                start_row = start_config['start_row']
                start_column = start_config['start_column']

                data_list: List[Any] = []

                # .xls ファイルの場合、max_column が取得できないため別途計算
                if filename.endswith('.xls'):
                    num_columns = max(len(row) for row in input_sheet.get_rows())
                else:
                    num_columns = input_sheet.max_column

                # 『人員集計表』の処理
                if filename == Staffingtable_filename:
                    # 処理年月に応じた列インデックスを算出
                    target_column_index = processing_year_month % 100 + 3  # 202405 -> 8, 202406 -> 9, 202407 -> 10

                    # データ行をループしてリストに追加
                    for row_index in range(start_row, start_row + 9):  # 59行目から67行目まで
                        segment = 'D' if row_index <= 61 else ('PD' if row_index <= 64 else 'F')
                        category = '社員' if row_index % 3 == 2 else ('引入外注者' if row_index % 3 == 1 else '有期契約社員')
                        personnel = input_sheet.cell(row=row_index, column=target_column_index).value

                        data_list.append({
                            '年月': processing_year_month,
                            'セグメント': segment,
                            '分類': category,
                            '人員': personnel
                        })

                    model = StaffingSummary  # 人員集計表のモデルを設定

                # 『受注売上受注残』の処理
                elif filename == SalesBacklog_filename:
                    # 処理年月に応じた列インデックスを算出 (202405 -> 9, 202406 -> 10, 202407 -> 11)
                    target_column_index = processing_year_month % 100 + 4

                    # 特定の行と列からデータを取得し、リストに追加
                    segments = ['国内D', '国内K', '国内S', '国内BD', '海外D', '海外K', '海外S', 'PD', 'IP', 'AW']
                    categories = ['受注高', '売上高', '受注残高']
                    start_rows = [6, 9, 12, 15, 21, 24, 27, 36, 39, 42]

                    for segment, start_row in zip(segments, start_rows):
                        for i, category in enumerate(categories):
                            amount = input_sheet.cell(row=start_row + i, column=target_column_index).value

                            data_list.append({
                                '年月': processing_year_month,
                                'セグメント': segment,
                                '科目': category,
                                '金額': amount
                            })

                    model = SalesBacklog

                # 『建仮計上額』の処理
                elif model is None and filename == ConstructionSuspenseAccounts_filename:
                    process_construction_suspense_accounts(input_excel_path, processing_year_month)
                    continue

                # 『原価差額調整計算表』の処理
                elif filename == CostVarianceAdjustment_filename:
                    process_cost_variance_adjustment(input_excel_path, processing_year_month)
                    continue

                # 『品目別在庫明細表』の処理
                elif filename == ItemizedInventoryDetails_filename:
                    process_itemized_inventory_details(input_excel_path, processing_year_month)
                    continue

                # 『PJ管理物件勘定内訳表』の処理
                elif filename == DirectExpenses_filename:
                    process_direct_expenses(input_excel_path, processing_year_month)
                    continue 

                # 『財務諸表』の処理
                elif filename == FinancialCostOfSales_filename:
                    process_financialcostOfsales(input_excel_path, processing_year_month)
                    continue 

                # その他のExcelファイルの処理
                else:
                    # 開始列を考慮してデータを取得
                    for row_index, row in enumerate(input_sheet.iter_rows(min_row=start_row, values_only=True)):
                        insert_data = {f"column{i+1}": row[i + start_column - 1] if i + start_column - 1 < len(row) else None for i in range(num_columns - start_column + 1)}
                        # ActualExpenses、Discounts、SalesReceivablesの場合、C列（"column3"）がNULLならスキップ
                        if model in (ActualExpenses, Discounts, SalesReceivables):
                            if insert_data.get("column3") is None:
                                continue  # このレコードは抽出対象外とする
                        data_list.append(insert_data)

                db = SessionLocal()
                try:
                    engine.echo = False
                    bulk_insert_with_chunk(db, model, data_list, chunk_size=1000)
                except Exception as e:
                    db.rollback()
                    logging.info(f"{filename}のデータ挿入中にエラーが発生しました：{e}")
                finally:
                    db.close()       
                
            progress = (index + 1) / total_files
            progress_bar.set(progress)
            root.update_idletasks()  # GUIを更新
            
        progress_bar.grid_remove()
    
    except ValueError:
        messagebox.showerror("エラー", "有効な年月を入力してください。")
    else:
        # 処理が正常に完了したことを通知
        messagebox.showinfo("処理完了", "処理が完了しました。")

def get_previous_month(year_month_str):
    """
    YYYYMM形式の文字列を受け取り、前月のYYYYMM形式の文字列を返す関数
    """
    year = int(year_month_str[:4])
    month = int(year_month_str[4:])

    # 現在の年月から1か月前に戻す
    if month == 1:
        year -= 1
        month = 12
    else:
        month -= 1

    # YYYYMM形式にフォーマット
    return f"{year}{month:02d}"

def reset_year_month():
    processing_year_month = entry_year_month.get()
    previous_month = get_previous_month(processing_year_month)
    
    if not check_year_month_format():
        return  # YYYYMM形式でない場合は処理を中断
    
    confirm = messagebox.askokcancel("確認", f"{processing_year_month}のデータをリセットします。よろしいですか？")
    if not confirm:
        return  # キャンセルされた場合は処理を中断

    # データベース接続を作成
    db = SessionLocal()
    
    #input_dir_path = 'C:/Users/1221050/Documents/work/01_業務/BI/Board/総務課/データソース/' + str(processing_year_month)
    input_dir_path = 'W:/WF_07_ｼｽﾃﾑ関係/16_BI/総務課/データソース/' + str(processing_year_month)
    #input_dir_path = 'C:/Users/KYOHEI/Documents/work/データソース/' + str(processing_year_month)
    
    logging.basicConfig(
        filename=input_dir_path + '/app.log',
        level=logging.DEBUG,
        format='%(asctime)s - %(levelname)s - %(message)s',
        filemode='a'
    )

    try:
        # 各テーブルと年月カラムを対応付ける
        tables = [
            (PurchaseCostVariance, PurchaseCostVariance.年月, processing_year_month),
            (CostCenterReports, CostCenterReports.column1, processing_year_month),  
            (ActualExpenses, ActualExpenses.column1, processing_year_month),  
            (SalesReceivables, SalesReceivables.column10, processing_year_month),  
            (Discounts, Discounts.column1, processing_year_month),  
            (StaffingSummary, StaffingSummary.年月, processing_year_month),
            (SalesBacklog, SalesBacklog.年月, processing_year_month),
            (ConstructionSuspenseAccounts, ConstructionSuspenseAccounts.年月, processing_year_month),
            (ItemizedInventoryDetails, ItemizedInventoryDetails.年月, processing_year_month),
            (DirectExpenses, DirectExpenses.年月, processing_year_month),
            (PurchasePriceGap, PurchasePriceGap.column1, previous_month),  # 前月を使用
            (ItemList, ItemList.column1, previous_month),  # 前月を使用
            (FinancialCostOfSales, FinancialCostOfSales.年月, processing_year_month),
            (FinancialOperatingIncome, FinancialOperatingIncome.年月, processing_year_month)
        ]

        # 各テーブルごとに削除処理を実行
        for table, year_month_column, target_year_month in tables:
            # 該当するデータが存在するかを確認
            existing_records = db.execute(select(table).where(year_month_column == target_year_month)).scalars().all()
            
            if existing_records:  # 削除対象が存在する場合のみ削除
                delete_stmt = delete(table).where(year_month_column == target_year_month)
                db.execute(delete_stmt)
                logging.info(f"{table.__tablename__}の{target_year_month}のレコードを削除しました。")
            else:
                logging.info(f"{table.__tablename__}に{target_year_month}の削除対象はありませんでした。")

        db.commit()  # 変更を保存
        messagebox.showinfo("リセット完了", f"{processing_year_month}のデータがリセットされました。")

    except Exception as e:
        db.rollback()  # エラー発生時はロールバック
        logging.error(f"データ削除中にエラーが発生しました: {e}")
        messagebox.showerror("エラー", "リセット中にエラーが発生しました。ログを確認してください。")

    finally:
        db.close()

# GUIの作成
root = ctk.CTk()
root.title("MonthlyReport")
root.geometry("500x300")  # ウィンドウサイズを大きく設定

# 行と列の重み付けを設定（中央寄せのために均等に空白を作る）
root.grid_rowconfigure(0, weight=1)  # 上の余白
root.grid_rowconfigure(1, weight=1)  # ラベル、エントリーフィールドのある行
root.grid_rowconfigure(2, weight=1)  # ボタンのある行
root.grid_rowconfigure(3, weight=1)  # 下の余白

root.grid_columnconfigure(0, weight=1)  # 左側の余白
root.grid_columnconfigure(1, weight=1)  # 右側の余白

# ラベルの作成
label = ctk.CTkLabel(root, text="処理年月 (YYYYMM) を入力してください:", font=("Meiryo UI", 16))
label.grid(row=1, column=0, columnspan=2, padx=20, pady=20, sticky="nsew")

# エントリー（入力フィールド）の作成
entry_year_month = ctk.CTkEntry(root, placeholder_text="例: 202401", width=200, height=40, font=("Meiryo UI", 14))
entry_year_month.grid(row=2, column=0, columnspan=2, padx=20, pady=10, sticky="")

# ボタンの作成
button = ctk.CTkButton(root, text="実行", command=process_year_month, width=120, height=40, font=("Meiryo UI", 16))
reset_button = ctk.CTkButton(root, text="リセット", command=reset_year_month, fg_color="red", width=120, height=40, font=("Meiryo UI", 16))
button.grid(row=3, column=0, padx=20, pady=30, sticky="e")
reset_button.grid(row=3, column=1, padx=20, pady=30, sticky="w")

progress_bar = ctk.CTkProgressBar(root, width=300, height=20)
progress_bar.set(0)  # 初期値を0に設定
progress_bar.grid(row=4, column=0, columnspan=2, padx=20, pady=20)  # 初期状態では非表示
progress_bar.grid_remove()  # 初期状態で非表示にする

root.mainloop()